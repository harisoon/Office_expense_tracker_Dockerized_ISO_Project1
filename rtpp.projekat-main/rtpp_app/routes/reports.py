from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from rtpp_app.extensions import db
from rtpp_app.models.categorybudgets import CategoryBudget
from rtpp_app.models.category import Category
from rtpp_app.models.vendor import Vendor
from rtpp_app.models.invoice import Invoice
from rtpp_app.models.invoiceitem import InvoiceItem
from datetime import datetime, date
from calendar import monthrange, month_name
import pandas as pd
import io
from sqlalchemy import and_, extract, func, or_

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

def get_available_years():

    invoice_years = db.session.query(
        extract('year', Invoice.created_at).label('year')
    ).distinct().subquery()
    

    budget_years = db.session.query(
        extract('year', CategoryBudget.created_at).label('year')
    ).distinct().subquery()
    

    all_years = db.session.query(
        func.coalesce(invoice_years.c.year, budget_years.c.year).label('year')
    ).select_from(
        invoice_years.join(budget_years, invoice_years.c.year == budget_years.c.year, full=True)
    ).distinct().order_by('year').all()
    
    if not all_years:

        current_year = datetime.now().year
        return [current_year - 1, current_year, current_year + 1]
    
    years = [year[0] for year in all_years if year[0] is not None]
    

    current_year = datetime.now().year
    if current_year not in years:
        years.append(current_year)
    
    return sorted(years)

@reports_bp.route('/')
@login_required
def reports_dashboard():
    """Main reports dashboard page"""
    current_date = datetime.now()
    current_month = current_date.month
    current_month_name = month_name[current_month]
    current_year = current_date.year
    
    # Get available years from data
    available_years = get_available_years()
    
    # Get available months/years from data
    available_dates = db.session.query(
        extract('year', CategoryBudget.created_at).label('year'),
        extract('month', CategoryBudget.created_at).label('month')
    ).distinct().order_by('year', 'month').all()
    
    return render_template('reports.html', 
                         current_month=current_month,
                         current_year=current_year,
                         current_month_name=current_month_name,
                         available_dates=available_dates,
                         available_years=available_years)

@reports_bp.route('/monthly-budget-report')
@login_required
def generate_monthly_budget_report():
    """Generate Excel report for monthly budgets"""
    
    # Get parameters from request
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    report_format = request.args.get('format', 'detailed')
    include_expenses = request.args.get('include_expenses') is not None
    include_vendors = request.args.get('include_vendors') is not None
    
    try:
        # Get budget data for the specified month/year
        budget_data = get_monthly_budget_data(month, year)
        
        if not budget_data:
            flash(f'No budget data found for {month}/{year}', 'warning')
            return redirect(url_for('reports.reports_dashboard'))
        
        # Create Excel file
        excel_file = create_budget_excel_report(
            budget_data, month, year, report_format, include_expenses, include_vendors
        )
        
        # Generate filename
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        filename = f"HTEC_Budget_Report_{month_names[month]}_{year}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for('reports.reports_dashboard'))

@reports_bp.route('/yearly-budget-report')
@login_required
def generate_yearly_budget_report():
    """Generate Excel report for yearly budgets"""
    
    # Get parameters from request
    year = request.args.get('year', datetime.now().year, type=int)
    report_type = request.args.get('report_type', 'summary')
    
    try:
        # Get yearly budget data
        yearly_data = get_yearly_budget_data(year, report_type)
        
        if not yearly_data:
            flash(f'No budget data found for year {year}', 'warning')
            return redirect(url_for('reports.reports_dashboard'))
        
        # Create Excel file
        excel_file = create_yearly_excel_report(yearly_data, year, report_type)
        
        # Generate filename
        filename = f"HTEC_Yearly_Budget_Report_{year}_{report_type}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating yearly report: {str(e)}', 'error')
        return redirect(url_for('reports.reports_dashboard'))

def get_monthly_budget_data(month, year):
    """Fetch budget data for a specific month/year"""
    
    # Get first and last day of the month
    first_day = date(year, month, 1)
    last_day_num = monthrange(year, month)[1]
    last_day = date(year, month, last_day_num)
    
    # Query budgets for the specified month
    budgets = CategoryBudget.query.filter(
        and_(
            CategoryBudget.created_at >= first_day,
            CategoryBudget.created_at <= last_day
        )
    ).all()
    
    # If no budgets found for exact month, get latest budgets
    if not budgets:
        budgets = CategoryBudget.query.filter(
            CategoryBudget.created_at <= last_day
        ).order_by(CategoryBudget.created_at.desc()).limit(10).all()
    
    # Prepare data structure
    budget_data = []
    total_budget = 0
    total_spent = 0
    
    for budget in budgets:
        # Get category-specific invoices for the month
        category_invoices = db.session.query(func.sum(Invoice.total_excl_vat)).filter(
            and_(
                Invoice.created_at >= first_day,
                Invoice.created_at <= last_day
            )
        ).scalar() or 0
        
        budget_amount = getattr(budget, 'total_value', getattr(budget, 'budget_amount', 0)) or 0
        spent_amount = getattr(budget, 'spent_value', getattr(budget, 'spent_amount', category_invoices)) or 0
        remaining = budget_amount - spent_amount
        utilization = (spent_amount / budget_amount * 100) if budget_amount > 0 else 0
        
        budget_data.append({
            'Category': budget.category.name if budget.category else 'Unknown',
            'Budgeted Amount (KM)': budget_amount,
            'Spent Amount (KM)': spent_amount,
            'Remaining (KM)': remaining,
            'Utilization (%)': round(utilization, 2),
            'Status': 'Over Budget' if spent_amount > budget_amount else 'Within Budget'
        })
        
        total_budget += budget_amount
        total_spent += spent_amount
    
    total_remaining = total_budget - total_spent
    total_utilization = (total_spent / total_budget * 100) if total_budget > 0 else 0
    
    budget_data.append({
        'Category': 'TOTAL',
        'Budgeted Amount (KM)': total_budget,
        'Spent Amount (KM)': total_spent,
        'Remaining (KM)': total_remaining,
        'Utilization (%)': round(total_utilization, 2),
        'Status': 'Summary'
    })
    
    return budget_data

def get_yearly_budget_data(year, report_type='summary'):
    """Fetch budget data for a specific year"""
    
    first_day = date(year, 1, 1)
    last_day = date(year, 12, 31)
    
    if report_type == 'monthly_breakdown':
        # Get data for each month
        yearly_data = []
        
        for month in range(1, 13):
            monthly_data = get_monthly_budget_data(month, year)
            if monthly_data:
                # Add month identifier to each row except totals
                month_name_str = month_name[month]
                for item in monthly_data[:-1]:  # Exclude total row
                    item['Month'] = month_name_str
                    yearly_data.append(item)
        
        return yearly_data
    
    else:
        # Get yearly summary data
        budgets = CategoryBudget.query.filter(
            and_(
                CategoryBudget.created_at >= first_day,
                CategoryBudget.created_at <= last_day
            )
        ).all()
        
        if not budgets:
            # Get latest budgets if no budgets for the year
            budgets = CategoryBudget.query.filter(
                CategoryBudget.created_at <= last_day
            ).order_by(CategoryBudget.created_at.desc()).limit(20).all()
        
        # Aggregate yearly data by category
        category_data = {}
        
        for budget in budgets:
            category_name = budget.category.name if budget.category else 'Unknown'
            
            if category_name not in category_data:
                category_data[category_name] = {
                    'total_value': 0,
                    'spent_value': 0
                }
            
            budget_amount = getattr(budget, 'total_value', getattr(budget, 'budget_amount', 0)) or 0
            spent_amount = getattr(budget, 'spent_value', getattr(budget, 'spent_amount', 0)) or 0
            
            category_data[category_name]['total_value'] += budget_amount
            category_data[category_name]['spent_value'] += spent_amount
        
        # Get yearly invoice totals by category if available
        yearly_invoices = db.session.query(func.sum(Invoice.total_excl_vat)).filter(
            and_(
                Invoice.created_at >= first_day,
                Invoice.created_at <= last_day
            )
        ).scalar() or 0
        
        # Convert to list format
        yearly_data = []
        total_budget = 0
        total_spent = 0
        
        for category_name, data in category_data.items():
            budget_amount = data['total_value']
            spent_amount = data['spent_value']
            remaining = budget_amount - spent_amount
            utilization = (spent_amount / budget_amount * 100) if budget_amount > 0 else 0
            
            yearly_data.append({
                'Category': category_name,
                'Budgeted Amount (KM)': budget_amount,
                'Spent Amount (KM)': spent_amount,
                'Remaining (KM)': remaining,
                'Utilization (%)': round(utilization, 2),
                'Status': 'Over Budget' if spent_amount > budget_amount else 'Within Budget'
            })
            
            total_budget += budget_amount
            total_spent += spent_amount
        
        # Add yearly summary row
        total_remaining = total_budget - total_spent
        total_utilization = (total_spent / total_budget * 100) if total_budget > 0 else 0
        
        yearly_data.append({
            'Category': 'YEARLY TOTAL',
            'Budgeted Amount (KM)': total_budget,
            'Spent Amount (KM)': total_spent,
            'Remaining (KM)': total_remaining,
            'Utilization (%)': round(total_utilization, 2),
            'Status': 'Summary'
        })
        
        return yearly_data

def create_budget_excel_report(budget_data, month, year, report_format='detailed', include_expenses=True, include_vendors=True):
    """Create Excel file with budget report"""
    
    # Create DataFrame
    df = pd.DataFrame(budget_data)
    
    # Create Excel writer object
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main budget data
        df.to_excel(writer, sheet_name='Budget Summary', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Budget Summary']
        
        # Format the worksheet
        format_budget_worksheet(workbook, worksheet, df, month, year)
        
        # Add additional sheets based on format and options
        if report_format in ['detailed', 'with-charts'] and include_expenses:
            add_detailed_expenses_sheet(writer, month, year)
        
        if include_vendors:
            add_vendor_breakdown_sheet(writer, month, year)
        
        if report_format == 'with-charts':
            add_charts_sheet(writer, budget_data, month, year)
    
    output.seek(0)
    return output

def create_yearly_excel_report(yearly_data, year, report_type='summary'):
    """Create Excel file for yearly report"""
    
    df = pd.DataFrame(yearly_data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main yearly data
        sheet_name = f'{year} Summary' if report_type != 'monthly_breakdown' else f'{year} Monthly Breakdown'
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Format the worksheet
        format_yearly_worksheet(workbook, worksheet, df, year, report_type)
        
        # Add additional analysis sheets
        add_yearly_analysis_sheet(writer, year)
        add_yearly_trends_sheet(writer, year)
    
    output.seek(0)
    return output

def format_budget_worksheet(workbook, worksheet, df, month, year):
    """Format the budget worksheet with professional styling"""
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    
    # Define color scheme
    colors = {
        'primary': '1E3A8A',      # Dark blue
        'secondary': '3B82F6',    # Blue
        'success': '10B981',      # Green
        'warning': 'F59E0B',      # Amber
        'danger': 'EF4444',       # Red
        'light_gray': 'F8FAFC',   # Very light gray
        'medium_gray': 'E2E8F0',  # Medium gray
        'dark_gray': '64748B'     # Dark gray
    }
    
    # Create border styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # Insert rows for header
    worksheet.insert_rows(1, 5)
    
    # Company header
    worksheet['A1'] = 'HTEC TUZLA d.o.o.'
    worksheet['A1'].font = Font(size=20, bold=True, color=colors['primary'])
    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells('A1:F1')
    worksheet.row_dimensions[1].height = 30
    
    # Report title
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
    title = f"Monthly Budget Report - {month_names[month]} {year}"
    worksheet['A2'] = title
    worksheet['A2'].font = Font(size=16, bold=True, color=colors['dark_gray'])
    worksheet['A2'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells('A2:F2')
    worksheet.row_dimensions[2].height = 25
    
    # Generation info
    worksheet['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    worksheet['A3'].font = Font(size=10, italic=True, color=colors['dark_gray'])
    worksheet['A3'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells('A3:F3')
    
    # Add empty row for spacing
    worksheet.row_dimensions[4].height = 15
    
    # Style headers (row 6 after insertions)
    header_fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    worksheet.row_dimensions[6].height = 20
    
    # Style data rows
    for row in range(7, len(df) + 7):
        row_height = 18
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
            
            # Highlight total row
            if worksheet.cell(row=row, column=1).value == 'TOTAL':
                cell.fill = PatternFill(start_color=colors['medium_gray'], end_color=colors['medium_gray'], fill_type='solid')
                cell.font = Font(bold=True, size=11)
                cell.border = thick_border
                row_height = 22
            
            # Format numbers
            if col > 1 and col < len(df.columns):  # Numeric columns
                if isinstance(cell.value, (int, float)):
                    if 'KM' in df.columns[col-1]:
                        cell.number_format = '#,##0.00'
                    elif '%' in df.columns[col-1]:
                        cell.number_format = '0.00%'
                        cell.value = cell.value / 100 if cell.value > 1 else cell.value
            
            # Color-code status column
            if col == len(df.columns):  
                if cell.value == 'Over Budget':
                    cell.fill = PatternFill(start_color=colors['danger'], end_color=colors['danger'], fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                elif cell.value == 'Within Budget':
                    cell.fill = PatternFill(start_color=colors['success'], end_color=colors['success'], fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                elif cell.value == 'Summary':
                    cell.fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
        
        worksheet.row_dimensions[row].height = row_height
    
    # Auto-adjust column widths with minimum and maximum constraints
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        column = get_column_letter(col)
        
        # Check header length
        header_length = len(str(df.columns[col-1]))
        max_length = max(max_length, header_length)
        
        # Check data lengths
        for row in range(7, len(df) + 7):
            try:
                cell_length = len(str(worksheet[f'{column}{row}'].value))
                max_length = max(max_length, cell_length)
            except:
                pass
        
        # Set width with constraints
        adjusted_width = max(12, min(max_length + 3, 25))
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Add conditional formatting for utilization percentage
    if len(df.columns) >= 5:  # Assuming utilization is 5th column
        util_col = get_column_letter(5)
        util_range = f"{util_col}7:{util_col}{len(df) + 6}"
        
        # Create color scale rule (red to green)
        color_scale = ColorScaleRule(
            start_type='percentile', start_value=0, start_color='FF6B6B',
            mid_type='percentile', mid_value=50, mid_color='FFD93D',
            end_type='percentile', end_value=100, end_color='51CF66'
        )
        worksheet.conditional_formatting.add(util_range, color_scale)

def format_yearly_worksheet(workbook, worksheet, df, year, report_type):
    """Format the yearly worksheet with professional styling"""
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define color scheme (same as monthly)
    colors = {
        'primary': '1E3A8A',
        'secondary': '3B82F6',
        'success': '10B981',
        'warning': 'F59E0B',
        'danger': 'EF4444',
        'light_gray': 'F8FAFC',
        'medium_gray': 'E2E8F0',
        'dark_gray': '64748B'
    }
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # Insert header rows
    worksheet.insert_rows(1, 5)
    
    # Company header
    worksheet['A1'] = 'HTEC TUZLA d.o.o.'
    worksheet['A1'].font = Font(size=20, bold=True, color=colors['primary'])
    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    end_col = get_column_letter(len(df.columns))
    worksheet.merge_cells(f'A1:{end_col}1')
    worksheet.row_dimensions[1].height = 30
    
    # Report title
    if report_type == 'monthly_breakdown':
        title = f"Yearly Budget Report (Monthly Breakdown) - {year}"
    else:
        title = f"Yearly Budget Report - {year}"
    
    worksheet['A2'] = title
    worksheet['A2'].font = Font(size=16, bold=True, color=colors['dark_gray'])
    worksheet['A2'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells(f'A2:{end_col}2')
    worksheet.row_dimensions[2].height = 25
    
    # Generation info
    worksheet['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    worksheet['A3'].font = Font(size=10, italic=True, color=colors['dark_gray'])
    worksheet['A3'].alignment = Alignment(horizontal='center')
    worksheet.merge_cells(f'A3:{end_col}3')
    
    # Style headers
    header_fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    worksheet.row_dimensions[6].height = 20
    
    # Style data rows
    for row in range(7, len(df) + 7):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
            
            # Highlight total/summary rows
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value and ('TOTAL' in str(cell_value).upper() or 'SUMMARY' in str(cell_value).upper()):
                cell.fill = PatternFill(start_color=colors['medium_gray'], end_color=colors['medium_gray'], fill_type='solid')
                cell.font = Font(bold=True)
                cell.border = thick_border
                worksheet.row_dimensions[row].height = 22
            else:
                worksheet.row_dimensions[row].height = 18
            
            # Format numbers
            if col > 1 and isinstance(cell.value, (int, float)):
                if 'KM' in str(df.columns[col-1]):
                    cell.number_format = '#,##0.00'
                elif '%' in str(df.columns[col-1]):
                    cell.number_format = '0.00%'
                    cell.value = cell.value / 100 if cell.value > 1 else cell.value
    
    # Auto-adjust column widths
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in range(1, len(df) + 7):
            try:
                cell_length = len(str(worksheet[f'{column}{row}'].value))
                max_length = max(max_length, cell_length)
            except:
                pass
        
        adjusted_width = max(12, min(max_length + 3, 25))
        worksheet.column_dimensions[column].width = adjusted_width

def add_detailed_expenses_sheet(writer, month, year):
    """Add detailed invoices breakdown sheet with better formatting"""
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl
    from openpyxl.chart import BarChart, Reference, PieChart, Series
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get first and last day of the month
    first_day = date(year, month, 1)
    last_day_num = monthrange(year, month)[1]
    last_day = date(year, month, last_day_num)
    
    # Query invoices for the month
    invoices = Invoice.query.filter(
        and_(
            Invoice.created_at >= first_day,
            Invoice.created_at <= last_day
        )
    ).order_by(Invoice.created_at.desc()).all()
    
    if invoices:
        invoice_data = []
        for inv in invoices:
            invoice_data.append({
                'Date': inv.created_at.strftime('%Y-%m-%d'),
                'Invoice Number': inv.invoice_number or 'Unknown',
                'Vendor': inv.vendor.vendor_name if inv.vendor else 'Unknown',
                'Amount Excl VAT (KM)': inv.total_excl_vat or 0,
                'VAT Amount (KM)': inv.vat_amount or 0,
                'Total Incl VAT (KM)': inv.total_incl_vat or 0,
                # removed old vat_rate column
                'Quantity': sum(item.quantity for item in inv.invoice_items or []),
            })
        
        df_invoices = pd.DataFrame(invoice_data)
        df_invoices.to_excel(writer, sheet_name='Detailed Invoices', index=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets['Detailed Invoices']
        
        # Insert header rows
        worksheet.insert_rows(1, 3)
        
        # Add title
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        worksheet['A1'] = f"Detailed Invoices - {month_names[month]} {year}"
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells(f'A1:{get_column_letter(len(df_invoices.columns))}1')
        
        # Format headers
        colors = {'primary': '1E3A8A'}
        header_fill = PatternFill(start_color=colors['primary'], end_color=colors['primary'], fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col in range(1, len(df_invoices.columns) + 1):
            cell = worksheet.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format data and add borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in range(5, len(df_invoices) + 5):
            for col in range(1, len(df_invoices.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if col > 3 else 'left')
                
                # Format currency columns
                if 'KM' in str(df_invoices.columns[col-1]):
                    cell.number_format = '#,##0.00'
                elif '%' in str(df_invoices.columns[col-1]):
                    cell.number_format = '0.00%'
        
        # Auto-adjust columns
        for col in range(1, len(df_invoices.columns) + 1):
            column = get_column_letter(col)
            max_length = max(len(str(df_invoices.columns[col-1])), 10)
            
            for row in range(5, len(df_invoices) + 5):
                try:
                    cell_length = len(str(worksheet[f'{column}{row}'].value))
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            worksheet.column_dimensions[column].width = min(max_length + 2, 30)

def add_vendor_breakdown_sheet(writer, month, year):
    """Add vendor breakdown analysis sheet"""
    
    # Get first and last day of the month
    first_day = date(year, month, 1)
    last_day_num = monthrange(year, month)[1]
    last_day = date(year, month, last_day_num)
    
    # Query vendor spending for the month
    vendor_spending = db.session.query(
        Vendor.vendor_name,
        func.count(Invoice.id).label('invoice_count'),
        func.sum(Invoice.total_excl_vat).label('total_spent'),
        func.avg(Invoice.total_excl_vat).label('avg_invoice'),
        func.max(Invoice.total_excl_vat).label('max_invoice')
    ).join(Invoice).filter(
        and_(
            Invoice.created_at >= first_day,
            Invoice.created_at <= last_day
        )
    ).group_by(Vendor.vendor_name).order_by(func.sum(Invoice.total_excl_vat).desc()).all()
    
    if vendor_spending:
        vendor_data = []
        for vendor in vendor_spending:
            vendor_data.append({
                'Vendor Name': vendor.vendor_name,
                'Number of Invoices': vendor.invoice_count,
                'Total Spent (KM)': round(vendor.total_spent or 0, 2),
                'Average Invoice (KM)': round(vendor.avg_invoice or 0, 2),
                'Largest Invoice (KM)': round(vendor.max_invoice or 0, 2)
            })
        
        df_vendors = pd.DataFrame(vendor_data)
        df_vendors.to_excel(writer, sheet_name='Vendor Breakdown', index=False)

def add_charts_sheet(writer, budget_data, month, year):
    """Add charts and visualizations sheet with improved formatting and chart objects"""

    # Prepare summary data for charts
    chart_data = []
    for item in budget_data[:-1]:  # Exclude total row
        chart_data.append({
            'Category': item['Category'],
            'Budgeted Amount (KM)': item['Budgeted Amount (KM)'],
            'Spent Amount (KM)': item['Spent Amount (KM)'],
            'Utilization (%)': item['Utilization (%)']
        })

    if not chart_data:
        return

    df_chart = pd.DataFrame(chart_data)
    sheet_name = 'Charts Data'
    df_chart.to_excel(writer, sheet_name=sheet_name, index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Style headers
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col in range(1, len(df_chart.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust columns
    for col in range(1, len(df_chart.columns) + 1):
        column = get_column_letter(col)
        max_length = max(len(str(df_chart.columns[col-1])), 10)
        for row in range(2, len(df_chart) + 2):
            try:
                cell_length = len(str(worksheet[f'{column}{row}'].value))
                max_length = max(max_length, cell_length)
            except Exception:
                pass
        worksheet.column_dimensions[column].width = min(max_length + 2, 30)

    # Add a bar chart for Budget vs Spent
    bar_chart = BarChart()
    bar_chart.title = f"Budget vs Spent by Category - {month_name[month]} {year}"
    bar_chart.y_axis.title = 'Amount (KM)'
    bar_chart.x_axis.title = 'Category'

    categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(df_chart) + 1)
    budget_data_ref = Reference(worksheet, min_col=2, min_row=1, max_row=len(df_chart) + 1)
    spent_data_ref = Reference(worksheet, min_col=3, min_row=1, max_row=len(df_chart) + 1)

    bar_chart.add_data(budget_data_ref, titles_from_data=True)
    bar_chart.add_data(spent_data_ref, titles_from_data=True)
    bar_chart.set_categories(categories)
    bar_chart.width = 20
    bar_chart.height = 10

    worksheet.add_chart(bar_chart, f"E2")

    # Add a pie chart for Spent distribution
    pie_chart = PieChart()
    pie_chart.title = f"Spent Distribution by Category - {month_name[month]} {year}"
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(df_chart) + 1)
    spent_values = Reference(worksheet, min_col=3, min_row=1, max_row=len(df_chart) + 1)
    pie_chart.add_data(spent_values, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.height = 8
    pie_chart.width = 8

    worksheet.add_chart(pie_chart, f"E20")

    # Add a bar chart for Utilization %
    util_chart = BarChart()
    util_chart.title = f"Utilization (%) by Category - {month_name[month]} {year}"
    util_chart.y_axis.title = 'Utilization (%)'
    util_chart.x_axis.title = 'Category'
    util_data_ref = Reference(worksheet, min_col=4, min_row=1, max_row=len(df_chart) + 1)
    util_chart.add_data(util_data_ref, titles_from_data=True)
    util_chart.set_categories(categories)
    util_chart.width = 20
    util_chart.height = 8

    worksheet.add_chart(util_chart, f"E35")

def add_yearly_analysis_sheet(writer, year):
    """Add yearly analysis and trends sheet"""
    
    # Get quarterly data for the year
    quarterly_data = []
    quarters = [
        {'q': 1, 'months': [1, 2, 3], 'name': 'Q1'},
        {'q': 2, 'months': [4, 5, 6], 'name': 'Q2'},
        {'q': 3, 'months': [7, 8, 9], 'name': 'Q3'},
        {'q': 4, 'months': [10, 11, 12], 'name': 'Q4'}
    ]
    
    for quarter in quarters:
        total_budget = 0
        total_spent = 0
        
        for month in quarter['months']:
            try:
                first_day = date(year, month, 1)
                last_day = date(year, month, monthrange(year, month)[1])
                
                # Get total invoices for the month
                monthly_spent = db.session.query(func.sum(Invoice.total_excl_vat)).filter(
                    and_(
                        Invoice.created_at >= first_day,
                        Invoice.created_at <= last_day
                    )
                ).scalar() or 0
                
                total_spent += monthly_spent
                
                # Get budgets for the month (approximate)
                monthly_budgets = db.session.query(func.sum(CategoryBudget.total_value)).filter(
                    and_(
                        CategoryBudget.created_at >= first_day,
                        CategoryBudget.created_at <= last_day
                    )
                ).scalar() or 0
                
                total_budget += monthly_budgets
                
            except Exception:
                continue
        
        quarterly_data.append({
            'Quarter': quarter['name'],
            'Total Budget (KM)': round(total_budget, 2),
            'Total Spent (KM)': round(total_spent, 2),
            'Utilization (%)': round((total_spent / total_budget * 100) if total_budget > 0 else 0, 2)
        })
    
    if quarterly_data:
        df_quarterly = pd.DataFrame(quarterly_data)
        df_quarterly.to_excel(writer, sheet_name='Quarterly Analysis', index=False)

def add_yearly_trends_sheet(writer, year):
    """Add yearly trends and comparison sheet"""
    
    # Get monthly trends for the year
    monthly_trends = []
    
    for month in range(1, 13):
        try:
            first_day = date(year, month, 1)
            last_day = date(year, month, monthrange(year, month)[1])
            
            # Get monthly statistics
            monthly_invoices = db.session.query(
                func.count(Invoice.id),
                func.sum(Invoice.total_excl_vat),
                func.avg(Invoice.total_excl_vat)
            ).filter(
                and_(
                    Invoice.created_at >= first_day,
                    Invoice.created_at <= last_day
                )
            ).first()
            
            monthly_trends.append({
                'Month': month_name[month],
                'Invoice Count': monthly_invoices[0] or 0,
                'Total Spending (KM)': round(monthly_invoices[1] or 0, 2),
                'Average Invoice (KM)': round(monthly_invoices[2] or 0, 2)
            })
            
        except Exception:
            monthly_trends.append({
                'Month': month_name[month],
                'Invoice Count': 0,
                'Total Spending (KM)': 0,
                'Average Invoice (KM)': 0
            })
    
    if monthly_trends:
        df_trends = pd.DataFrame(monthly_trends)
        df_trends.to_excel(writer, sheet_name='Monthly Trends', index=False)

@reports_bp.route('/quarterly-report')
@login_required
def generate_quarterly_report():
    """Generate quarterly budget report"""
    
    quarter = request.args.get('quarter', 1, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    
    # Calculate months for the quarter
    quarter_months = {
        1: [1, 2, 3],
        2: [4, 5, 6],
        3: [7, 8, 9],
        4: [10, 11, 12]
    }
    
    months = quarter_months.get(quarter, [1, 2, 3])
    
    # Get budget data for all months in quarter
    quarterly_data = []
    for month in months:
        monthly_data = get_monthly_budget_data(month, year)
        if monthly_data:
            # Add month identifier to each row except totals
            month_name_str = month_name[month]
            for item in monthly_data[:-1]:  # Exclude total row
                item['Month'] = month_name_str
                quarterly_data.append(item)
    
    if not quarterly_data:
        flash(f'No data found for Q{quarter} {year}', 'warning')
        return redirect(url_for('reports.reports_dashboard'))
    
    # Create quarterly Excel report
    excel_file = create_quarterly_excel_report(quarterly_data, quarter, year)
    filename = f"HTEC_Quarterly_Report_Q{quarter}_{year}.xlsx"
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def create_quarterly_excel_report(quarterly_data, quarter, year):
    """Create Excel file for quarterly report"""
    
    df = pd.DataFrame(quarterly_data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'Q{quarter} {year} Summary', index=False)
        
        # Add quarterly analysis
        add_quarterly_analysis(writer, quarter, year)
    
    output.seek(0)
    return output

def add_quarterly_analysis(writer, quarter, year):
    """Add quarterly analysis and summary"""
    
    quarter_months = {
        1: [1, 2, 3],
        2: [4, 5, 6],
        3: [7, 8, 9],
        4: [10, 11, 12]
    }
    
    months = quarter_months.get(quarter, [1, 2, 3])
    analysis_data = []
    
    for month in months:
        first_day = date(year, month, 1)
        last_day = date(year, month, monthrange(year, month)[1])
        
        # Get monthly statistics
        monthly_stats = db.session.query(
            func.count(Invoice.id),
            func.sum(Invoice.total_excl_vat)
        ).filter(
            and_(
                Invoice.created_at >= first_day,
                Invoice.created_at <= last_day
            )
        ).first()
        
        analysis_data.append({
            'Month': month_name[month],
            'Invoice Count': monthly_stats[0] or 0,
            'Total Spending (KM)': round(monthly_stats[1] or 0, 2)
        })
    
    if analysis_data:
        df_analysis = pd.DataFrame(analysis_data)
        df_analysis.to_excel(writer, sheet_name='Quarterly Analysis', index=False)

@reports_bp.route('/export-all-data')
@login_required
def export_all_data():
    """Export all budget and invoice data"""
    
    try:
        # Create comprehensive data export
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Export all budgets
            budgets = CategoryBudget.query.all()
            if budgets:
                budget_export = []
                for budget in budgets:
                    budget_export.append({
                        'ID': budget.id,
                        'Category': budget.category.name if budget.category else 'Unknown',
                        'Budget Amount': getattr(budget, 'total_value', getattr(budget, 'budget_amount', 0)),
                        'Spent Amount': getattr(budget, 'spent_value', getattr(budget, 'spent_amount', 0)),
                        'Created Date': budget.created_at.strftime('%Y-%m-%d') if budget.created_at else ''
                    })
                
                df_budgets = pd.DataFrame(budget_export)
                df_budgets.to_excel(writer, sheet_name='All Budgets', index=False)
            
            # Export all invoices
            invoices = Invoice.query.all()
            if invoices:
                invoice_export = []
            for inv in invoices:
                invoice_export.append({
                    'ID': inv.id,
                    'Invoice Number': inv.invoice_number,
                    'Vendor': inv.vendor.vendor_name if inv.vendor else 'Unknown',
                    'Total Excl VAT': inv.total_excl_vat,
                    'VAT Amount': inv.vat_amount,
                    'Total Incl VAT': inv.total_incl_vat,
                    'Created Date': inv.created_at.strftime('%Y-%m-%d') if inv.created_at else ''
                })
                
                df_invoices = pd.DataFrame(invoice_export)
                df_invoices.to_excel(writer, sheet_name='All Invoices', index=False)
            
            # Export all vendors
            vendors = Vendor.query.all()
            if vendors:
                vendor_export = []
                for vendor in vendors:
                    vendor_export.append({
                    'ID': vendor.id,
                    'Vendor Name': vendor.vendor_name,
                    'VAT Number': vendor.vat_number,
                    'Account Number': vendor.account_number,
                    'Bank': vendor.bank,
                    'Address': vendor.adress,
                    'City': vendor.vendorCity
                })
                
                df_vendors = pd.DataFrame(vendor_export)
                df_vendors.to_excel(writer, sheet_name='All Vendors', index=False)
        
        output.seek(0)
        filename = f"HTEC_Complete_Data_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('reports.reports_dashboard'))
